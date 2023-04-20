import os
import pandas as pd
from typing import Optional, Tuple, List
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.table import Table

from app.models.ASFT_Data import ASFT_Data

TEMPLATE_PATH = "C:/Users/lucas/Desktop/AA2000/app/templates/database_template.xlsx"


def load_excel_workbook(file_path: str, template_path: Optional[str] = None) -> Workbook:
    """
    Load an Excel workbook or a template file if the original file is not found.

    Args:
        file_path (str): The path to the Excel file to be loaded.
        template_path (Optional[str]): The path to a template Excel file to be loaded if the original file is not found.
                                        Default is None.

    Returns:
        Workbook: The loaded Workbook object.

    Raises:
        FileNotFoundError: If neither the file_path nor the template_path are found.
    """
    if os.path.exists(file_path):
        return load_workbook(file_path)
    if template_path and os.path.exists(template_path):
        return load_workbook(template_path)
    raise FileNotFoundError("Neither the file nor the template was found.")


def find_table(worksheet: Worksheet, table_name: str) -> Table:
    """
    Find and return the table with the specified name in the worksheet.

    Args:
        worksheet (Worksheet): The openpyxl Worksheet object containing the table.
        table_name (str): The name of the table to be found.

    Returns:
        Table: The found Table object.

    Raises:
        ValueError: If the table with the specified name is not found in the worksheet.
    """
    for tbl in worksheet.tables.values():
        if tbl.displayName == table_name:
            return tbl
    raise ValueError(f"The table '{table_name}' was not found in the worksheet.")


def add_dataframe_to_table(worksheet: Worksheet, table_name: str, data: pd.DataFrame) -> None:
    """
    Add the contents of a Pandas DataFrame to an existing table in an Excel worksheet.

    Args:
        worksheet (Worksheet): The openpyxl Worksheet object containing the table.
        table_name (str): The name of the table to which the DataFrame will be added.
        data (pd.DataFrame): The Pandas DataFrame containing the data to be added to the table.

    Returns:
        None
    """
    table = find_table(worksheet, table_name)

    table_range = CellRange(table.ref)
    row_start = table_range.max_row + 1
    row_end = row_start + len(data) - 1

    for row_idx, row_data in enumerate(data.values, start=row_start):
        for col_idx, cell_value in enumerate(row_data, start=table_range.min_col):
            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

    table.ref = f"A1:{worksheet.cell(row=row_end, column=table_range.max_col).coordinate}"


def is_key_present_in_table(
    worksheet: Worksheet, table_name: str, key_value: str, table_column_name: str = "key"
) -> bool:
    """
    Check if a key value is present in a specified column of a table within a worksheet.

    Args:
        worksheet (Worksheet): The openpyxl Worksheet object containing the table.
        table_name (str): The name of the table to search in.
        key_value (str): The key value to search for in the column.
        column_name (str, optional): The header of the column in which the key will be searched for. Defaults to "key".

    Raises:
        ValueError: If the specified column is not found in the table.

    Returns:
        bool: True if the key value is found in the column, False otherwise.
    """
    table = find_table(worksheet, table_name)
    table_range = CellRange(table.ref)

    column_idx = None
    for idx, cell in enumerate(worksheet[table_range.min_row]):
        if cell.value == table_column_name:
            column_idx = idx + 1
            break

    if column_idx is None:
        raise ValueError(f"The column with header '{table_column_name}' was not found in the table.")

    for row in range(table_range.min_row + 1, table_range.max_row + 1):
        cell_value = worksheet.cell(row=row, column=column_idx).value
        if cell_value == key_value:
            return True
    return False


def information_table(data: ASFT_Data) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "key": [data.key],
            "date": [data.date],
            "iata": [data.iata],
            "side": [data.side],
            "separation": [data.separation],
            "runway": [data.runway],
            "numbering": [data.numbering],
            "average speed": [data.average_speed],
            "fric_A": [data.fric_A],
            "fric_B": [data.fric_B],
            "fric_C": [data.fric_C],
            "org type": [data.org_type],
            "equipment": [data.equipment],
            "pilot": [data.pilot],
            "ice level": [data.ice_level],
            "runway length": [data.runway_length],
            "location": [data.location],
            "tyre type": [data.tyre_type],
            "tyre pressure": [data.tyre_pressure],
            "water film": [data.water_film],
            "system distance": [data.system_distance],
        }
    )


def database():
    wb = load_excel_workbook("C:/Users/lucas/Desktop/AA2000/db.xlsx", TEMPLATE_PATH)
    information_ws: Worksheet = wb["Information"]
    is_key_present_in_table(information_ws, "InformationTable", "2304181153EZE17L5")


if __name__ == "__main__":
    database()
