import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.table import Table

from app.models.ASFT_Data import ASFT_Data
from app.utils.calculations import chainage_table


def load_excel_workbook(file_path: str) -> Workbook:
    """
    Load an Excel workbook from a file path.

    Args:
        file_path (str): The file path of the Excel workbook to load.

    Returns:
        Workbook: The openpyxl Workbook object representing the loaded workbook.

    Raises:
        FileNotFoundError: If the specified file does not exist.
    """
    if os.path.exists(file_path):
        return load_workbook(file_path)
    raise FileNotFoundError("File not found.")


def find_table(worksheet: Worksheet, table_name: str) -> Table:
    """
    Find and return the table with the specified name in the worksheet.

    Args:
        worksheet (Worksheet): The openpyxl Worksheet object containing the table.
        table_name (str): The name of the table to be found.

    Returns:
        Table: The found Table object.

    Raises:
        ValueError: If the table with the specified name is not found in the worksheet.
    """
    for table in worksheet.tables.values():
        if table.displayName == table_name:
            return table
    raise ValueError(f"The table '{table_name}' was not found in the worksheet.")


def add_dataframe_to_table(worksheet: Worksheet, table_name: str, data: pd.DataFrame) -> None:
    """
    Add the contents of a Pandas DataFrame to an existing table in an Excel worksheet.

    Args:
        worksheet (Worksheet): The openpyxl Worksheet object containing the table.
        table_name (str): The name of the table to which the DataFrame will be added.
        data (pd.DataFrame): The Pandas DataFrame containing the data to be added to the table.

    Returns:
        None
    """
    table = find_table(worksheet, table_name)
    table_range = CellRange(table.ref)

    for row_idx, row_data in enumerate(data.values, start=table_range.max_row + 1):
        for col_idx, cell_value in enumerate(row_data, start=table_range.min_col):
            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

    table.ref = f"{worksheet.cell(row=table_range.min_row, column=table_range.min_col).coordinate}:{worksheet.cell(row=table_range.max_row + len(data), column=table_range.max_col).coordinate}"


def is_key_present_in_table(
    worksheet: Worksheet, table_name: str, key_value: str, table_column_name: str = "key"
) -> bool:
    """
    Check if a key value is present in a specified column of a table within a worksheet.

    Args:
        worksheet (Worksheet): The openpyxl Worksheet object containing the table.
        table_name (str): The name of the table to search in.
        key_value (str): The key value to search for in the column.
        column_name (str, optional): The header of the column in which the key will be searched for. Defaults to "key".

    Raises:
        ValueError: If the specified column is not found in the table.

    Returns:
        bool: True if the key value is found in the column, False otherwise.
    """
    table = find_table(worksheet, table_name)
    table_range = CellRange(table.ref)

    column_idx = None
    for idx, cell in enumerate(worksheet[table_range.min_row]):
        if cell.value == table_column_name:
            column_idx = idx + 1
            break

    if column_idx is None:
        raise ValueError(f"The column with header '{table_column_name}' was not found in the table.")

    for row in range(table_range.min_row + 1, table_range.max_row + 1):
        cell_value = worksheet.cell(row=row, column=column_idx).value
        if cell_value == key_value:
            return True
    return False


def information_table(data: ASFT_Data) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "key": [data.key],
            "date": [data.date],
            "iata": [data.iata],
            "side": [data.side],
            "separation": [data.separation],
            "runway": [data.runway],
            "numbering": [data.numbering],
            "average speed": [data.average_speed],
            "fric_A": [data.fric_A],
            "fric_B": [data.fric_B],
            "fric_C": [data.fric_C],
            "org type": [data.org_type],
            "equipment": [data.equipment],
            "pilot": [data.pilot],
            "ice level": [data.ice_level],
            "location": [data.location],
            "tyre type": [data.tyre_type],
            "tyre pressure": [data.tyre_pressure],
            "water film": [data.water_film],
            "system distance": [data.system_distance],
        }
    )


def measurements_table(data: ASFT_Data, runway_length: int, starting_point: int) -> pd.DataFrame:
    """
    Aligns the measurements table with the corresponding chainage of the runway, measured from left to right.

    This function takes runway numbering and runway length as inputs, calculates the chainage table, and then aligns
    the measurements data with the corresponding chainage values based on the starting point. The chainage values are
    measured from left to right, starting from the runway numbers that are between 01 and 18. The resulting DataFrame
    contains the Key, chainage, and measurements columns.

    Args:
        data (ASFT_Data): An instance of the ASFT_Data class, which contains the runway numbering, key, and measurements.
        runway_length (int): The total length of the runway, which should be a positive integer value.
        starting_point (int): The chainage value where the measurements data should start aligning, referenced from the
                              runway numbers between 01 and 18.

    Returns:
        pd.DataFrame: A pandas DataFrame containing the Key, chainage, and measurements columns, where the measurements
        data is aligned with the corresponding chainage values based on the starting point.

    Raises:
        ValueError: If the measurements table overflows the chainage table. This error suggests adjusting the starting
                    point or the runway length.


        |=============|====================================================================|=============|
        | -> -> -> -> |11   ===   ===   ===   ===   [ RUNWAY ]   ===   ===   ===   ===   29| <- <- <- <- |
        |=============|====================================================================|=============|

        .................................................................................................. chainage
        [ 0 ]                                                                                   [ LENGTH ]


                      .................................................................................... starting point from header 11
                      [ START ]  -> -> -> -> -> -> -> -> -> -> -> -> -> ->-> -> -> -> -> -> -> -> -> -> ->


                                                                                           ............... starting point from header 29
        <- <- <- <- <- <- <- <- <- <- <- <- <- <- <- <- <- <- <- <- <- <- <- <- <- <- <-   [ START ]

    """
    numbering = int(data.numbering)
    reverse = True if 19 <= numbering <= 36 else False if 1 <= numbering <= 18 else None

    chainage = chainage_table(runway_length, reversed=reverse)
    measurements = data.measurements

    start_index = chainage[chainage["chainage"] == starting_point].index[0]

    if start_index + len(measurements) > len(chainage):
        raise ValueError(
            "The measurements table overflows the chainage table. Please adjust the starting point or the runway length."
        )

    for col in measurements.columns:
        if col not in chainage.columns:
            chainage[col] = 0

        for i, value in enumerate(measurements[col]):
            chainage.at[start_index + i, col] = value

    chainage.insert(0, "Key", data.key)
    return chainage


data = ASFT_Data("C:/Users/lucas/Desktop/AA2000/data/EZE/EZE RWY 35 L3_230317_143552.pdf")
md = measurements_table(data, 3110, 3000)

print(md.head(40))
print(md.tail(40))


# TODO: measurement table for each side. adjust chianage depending on orientation
