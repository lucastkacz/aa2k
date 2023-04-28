from app.models.ASFT_Data import ASFT_Data
from typing import List, Tuple, Any
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
import pandas as pd

# TODO: fix friction_thirds (rename and fix types) and friction_interval_mean (rename and fix types)


def validate_lengths(L: Any, R: Any) -> None:
    """
    Validates that the lengths of L and R objects are the same.

    Args:
        L (Any): An object with a length (e.g., list, tuple, string).
        R (Any): An object with a length (e.g., list, tuple, string).

    Raises:
        ValueError: If the lengths of L and R are not the same.
    """
    if not hasattr(L, "__len__") or not hasattr(R, "__len__"):
        raise TypeError("Both input objects must have a length attribute.")

    if len(L) != len(R):
        raise ValueError(f"Lengths of L and R must be the same. Found {len(L)} and {len(R)}")


def validate_attributes(L: ASFT_Data, R: ASFT_Data, attributes: List[str]) -> None:
    """
    Validates that specified attributes are the same for both ASFT_Data objects and that side attributes are correct.

    Args:
        L (ASFT_Data): ASFT_Data object with side 'L'.
        R (ASFT_Data): ASFT_Data object with side 'R'.
        attributes (List[str]): List of attribute names to compare.

    Raises:
        ValueError: If side attributes are incorrect or specified attributes are not the same for both objects.
    """
    if L.side != "L":
        raise ValueError(f"Side attribute for L must be 'L'. Found '{L.side}'")

    if R.side != "R":
        raise ValueError(f"Side attribute for R must be 'R'. Found '{R.side}'")

    for attribute in attributes:
        value1 = getattr(L, attribute)
        value2 = getattr(R, attribute)

        if value1 != value2:
            raise ValueError(f"{attribute} values must be the same for both objects. Found '{value1}' and '{value2}'")


def setup_workbook(template_path: str, title: str) -> Tuple[Workbook, Worksheet]:
    """
    Sets up the workbook using the provided template.

    Args:
        template_path (str): Path to the template file.
        L (ASFT_Data): ASFT_Data object with side 'L'.

    Returns:
        Tuple[Workbook, Worksheet]: Loaded workbook and active worksheet.
    """
    wb = load_workbook(template_path)
    ws = wb.active
    ws.title = title
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.scale = 95

    return wb, ws


def friction_thirds(data: ASFT_Data) -> List[float]:
    length: int = len(data)
    third: float = length / 3
    thirds = []
    for row in range(length):
        if row <= round(third - 1):
            thirds.append(data.fric_A)
        elif row <= round(2 * third - 1):
            thirds.append(data.fric_B)
        else:
            thirds.append(data.fric_C)
    return thirds


def friction_interval_mean(df: pd.DataFrame, values: str, interval: int = 10) -> pd.Series:
    def get_friction_mean(series: pd.Series) -> pd.Series:
        fm: pd.Series = series.rolling(window=interval).mean()
        return fm[interval - 1 :: interval]

    return get_friction_mean(df[values]).reindex(range(len(df))).bfill()


def center_and_bold_cells(ws: Worksheet, min_row: int, min_col: int) -> None:
    # TODO: add a max_row and max_col or a range
    border = Border(
        left=Side(border_style="medium", color="000000"),
        right=Side(border_style="medium", color="000000"),
        top=Side(border_style="medium", color="000000"),
        bottom=Side(border_style="medium", color="000000"),
    )
    for row in ws.iter_rows(min_row=min_row, min_col=min_col):
        if row[0].row == 1:
            continue
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")


def get_file_name(data: ASFT_Data) -> str:
    return f"{data.iata} RWY{data.numbering} {data.separation}m {data.date.date()}"
