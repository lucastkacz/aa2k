import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.table import Table
from openpyxl.utils import column_index_from_string
from typing import Optional, Union
from pathlib import Path


def load_excel_workbook(file_path: str) -> Workbook:
    """
    Load an Excel workbook from a file path.

    Args:
        file_path (str): The file path of the Excel workbook to load.

    Returns:
        Workbook: The openpyxl Workbook object representing the loaded workbook.

    Raises:
        FileNotFoundError: If the specified file does not exist.
    """
    if os.path.exists(file_path):
        return load_workbook(file_path)
    raise FileNotFoundError("File not found.")


def find_table(worksheet: Worksheet, table_name: str) -> Table:
    """
    Find and return the table with the specified name in the worksheet.

    Args:
        worksheet (Worksheet): The openpyxl Worksheet object containing the table.
        table_name (str): The name of the table to be found.

    Returns:
        Table: The found Table object.

    Raises:
        ValueError: If the table with the specified name is not found in the worksheet.
    """
    for table in worksheet.tables.values():
        if table.displayName == table_name:
            return table
    raise ValueError(f"The table '{table_name}' was not found in the worksheet.")


def add_dataframe_to_table(worksheet: Worksheet, table_name: str, data: pd.DataFrame) -> None:
    """
    Add the contents of a Pandas DataFrame to an existing table in an Excel worksheet.

    Args:
        worksheet (Worksheet): The openpyxl Worksheet object containing the table.
        table_name (str): The name of the table to which the DataFrame will be added.
        data (pd.DataFrame): The Pandas DataFrame containing the data to be added to the table.

    Returns:
        None
    """
    table = find_table(worksheet, table_name)
    table_range = CellRange(table.ref)

    for row_idx, row_data in enumerate(data.values, start=table_range.max_row + 1):
        for col_idx, cell_value in enumerate(row_data, start=table_range.min_col):
            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

    table.ref = f"{worksheet.cell(row=table_range.min_row, column=table_range.min_col).coordinate}:{worksheet.cell(row=table_range.max_row + len(data), column=table_range.max_col).coordinate}"


def is_key_present_in_table(
    worksheet: Worksheet, table_name: str, key_value: str, table_column_name: str = "key"
) -> bool:
    """
    Check if a key value is present in a specified column of a table within a worksheet.

    Args:
        worksheet (Worksheet): The openpyxl Worksheet object containing the table.
        table_name (str): The name of the table to search in.
        key_value (str): The key value to search for in the column.
        column_name (str, optional): The header of the column in which the key will be searched for. Defaults to "key".

    Raises:
        ValueError: If the specified column is not found in the table.

    Returns:
        bool: True if the key value is found in the column, False otherwise.
    """
    table = find_table(worksheet, table_name)
    table_range = CellRange(table.ref)

    column_idx = None
    for idx, cell in enumerate(worksheet[table_range.min_row]):
        if cell.value == table_column_name:
            column_idx = idx + 1
            break

    if column_idx is None:
        raise ValueError(f"The column with header '{table_column_name}' was not found in the table.")

    for row in range(table_range.min_row + 1, table_range.max_row + 1):
        cell_value = worksheet.cell(row=row, column=column_idx).value
        if cell_value == key_value:
            return True
    return False


def write_column_to_excel(
    worksheet: Worksheet,
    start_row: int,
    column: str,
    data: pd.Series,
    format: Optional[Union[str, dict]] = None,
) -> None:
    """
    Writes a pandas Series to an Excel worksheet column with optional formatting.

    Args:
        worksheet (Worksheet): The Openpyxl Worksheet object where the data will be written.
        start_row (int): The row index at which the data writing should start (1-based indexing).
        column (str): The column name (e.g. 'A', 'B', 'C', etc.) where the data should be written.
        data (pd.Series): The pandas Series containing the data to be written to the Excel worksheet.
        format (Optional[Union[str, dict]], optional): The formatting to be applied to the cells in the column. This can
            be a string for number formatting or a dictionary for multiple attributes. Defaults to None.

    Examples:
        If you want to write a pandas Series with numbers to column 'A' starting at row 1 with a specific number format:
            write_column_to_excel(data, 1, 'A', worksheet, format='0.00')

        If you want to write a pandas Series with strings to column 'B' starting at row 1 with a specific font color:
            write_column_to_excel(data, 1, 'B', worksheet, format={'font': Font(color='FF0000')})

    Returns:
        None
    """
    column_index: int = column_index_from_string(column)

    for i, value in enumerate(data, start=start_row):
        cell = worksheet.cell(row=i, column=column_index)

        if format:
            if isinstance(format, str):
                cell.value = float(value)
                cell.number_format = format
            elif isinstance(format, dict):
                cell.value = value
                for attr, val in format.items():
                    if hasattr(cell, attr):
                        setattr(cell, attr, val)
        else:
            cell.value = value


def save_workbook(wb: Workbook, filename: str, output_folder: str) -> None:
    """
    Saves the workbook to the specified output folder.

    Args:
        wb (Workbook): Workbook to save.
        L (ASFT_Data): ASFT_Data object with side 'L'.
        output_folder (str): Path to the output folder.
    """
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    output_path = os.path.join(output_folder, filename)
    wb.save(output_path)


def merge_columns_into_thirds(number_of_rows: int, col: str, ws: Worksheet, start_row) -> pd.Series:
    """
    Merges cells in the given column into thirds.

    Args:
        number_of_rows (int): Length of the data or the amount of rows.
        col (str): Column identifier.
        ws (Worksheet): The target worksheet.
        start_row (int): The starting row for merging cells.
    """
    third: float = number_of_rows / 3
    row_A: int = start_row
    row_B: int = round(third) + start_row
    row_C: int = round(2 * third) + start_row
    end_row: int = number_of_rows + start_row
    ws.merge_cells(f"{col}{row_A}:{col}{row_B - 1}")
    ws.merge_cells(f"{col}{row_B}:{col}{row_C - 1}")
    ws.merge_cells(f"{col}{row_C}:{col}{end_row - 1}")


def merge_rows_in_range(length: int, col: str, ws: Worksheet, start_row: int, merge_range: int) -> None:
    """
    Merges cells vertically in the specified column to create average rows.

    Args:
        length (int): The total number of rows to merge.
        col (str): The column in which the cells should be merged.
        ws (Worksheet): The target worksheet.
        start_row (int): The starting row number for the merging process.
        merge_range (int): The number of consecutive rows to merge in each step.
    """
    for row in range(start_row, length + start_row, merge_range):
        ws.merge_cells(f"{col}{row}:{col}{row + merge_range - 1}")


def append_dataframe_to_excel(dataframe: pd.DataFrame, excel_file: Union[str, Path], sheet_name: str) -> None:
    file_path = Path(excel_file)

    if not file_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(list(dataframe.columns))
        for index, row in dataframe.iterrows():
            row_list = list(row)
            ws.append(row_list)
        wb.save(excel_file)
        return

    else:
        wb = load_workbook(file_path)
        if sheet_name in wb:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(list(dataframe.columns))

    for index, row in dataframe.iterrows():
        row_list = list(row)
        ws.append(row_list)

    wb.save(excel_file)
