import os.path
import openpyxl
import pandas as pd

def update_table(df: pd.DataFrame, db_path: str, table_name: str) -> None:
    if os.path.isfile(db_path):
        
        workbook = openpyxl.load_workbook(db_path)
        worksheet = workbook.active

        last_row = worksheet.max_row
        for row in df.values.tolist():
            worksheet.append(row)
        
        table = worksheet.tables[table_name]
        new_range = f"A1:{openpyxl.utils.get_column_letter(df.shape[1])}{last_row+len(df)}"
        table.ref = new_range

    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        headers = list(df.columns)
        worksheet.append(headers)
        for row in df.values.tolist():
            worksheet.append(row)
        
        table_range = f"A1:{openpyxl.utils.get_column_letter(df.shape[1])}{df.shape[0]+1}"
        table = openpyxl.worksheet.table.Table(displayName=table_name, ref=table_range)
        style = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        worksheet.add_table(table)

    workbook.save(db_path)

# TODO: avoid duplicate data
# TODO: unificar tablas