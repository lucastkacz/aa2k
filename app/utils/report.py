from app.models.ASFT_Data import ASFT_Data
from app.utils.functions.excel_functions import (
    write_column_to_excel,
    merge_columns_into_thirds,
    merge_rows_in_range,
)
from app.utils.functions.report_functions import (
    validate_attributes,
    validate_lengths,
    setup_workbook,
    friction_thirds,
    friction_interval_mean,
    center_and_bold_cells,
    get_file_name,
)

from pathlib import Path


from openpyxl.worksheet.worksheet import Worksheet

from typing import NewType

IntMultipleOf10 = NewType("PositiveIntMultipleOf10", int)


START_ROW = 11
START_COL = 2
MERGE_RANGE = 10


def populate_header_data(L: ASFT_Data, R: ASFT_Data, ws: Worksheet) -> None:
    """
    Populates the header information in the worksheet.

    Args:
        L (ASFT_Data): ASFT_Data object with side 'L'.
        R (ASFT_Data): ASFT_Data object with side 'R'.
        estado (str): The current state of the runway.
        pavimento (str): The type of pavement.
        ws (Worksheet): The target worksheet.
    """
    ws["C3"] = L.iata
    ws["E3"] = L.runway
    ws["G3"] = L.numbering
    ws["I3"] = f"{L.separation} m"
    ws["D4"] = L.date.date()
    ws["I5"] = L.equipment[-3:]
    ws["E6"] = int(L.average_speed)
    ws["H6"] = L.tyre_type
    ws["E7"] = L.weather
    ws["H7"] = L.runway_material
    ws["E8"] = L.date.time()
    ws["H8"] = R.date.time()


def write_report_no_chainage(L: ASFT_Data, R: ASFT_Data, output_folder: str) -> None:
    validate_attributes(L, R, ["iata", "runway", "numbering", "separation", "equipment", "tyre_type"])
    validate_lengths(L, R)

    L.measurements["Average Friction 100m"] = friction_interval_mean(L.measurements, "Friction")
    L.measurements["Thirds"] = friction_thirds(L)
    R.measurements["Average Friction 100m"] = friction_interval_mean(R.measurements, "Friction")
    R.measurements["Thirds"] = friction_thirds(R)

    name = get_file_name(L)

    wb, ws = setup_workbook("app/templates/report_template_without_chainage.xlsx", name)

    populate_header_data(L, R, ws)

    write_column_to_excel(ws, START_ROW, "B", L.measurements["Distance"], format="General")
    write_column_to_excel(ws, START_ROW, "C", L.measurements["Friction"], format="0.00")
    write_column_to_excel(ws, START_ROW, "D", L.measurements["Average Friction 100m"], format="0.00")
    write_column_to_excel(ws, START_ROW, "E", L.measurements["Thirds"], format="0.00")
    write_column_to_excel(ws, START_ROW, "F", R.measurements["Distance"], format="General")
    write_column_to_excel(ws, START_ROW, "G", R.measurements["Friction"], format="0.00")
    write_column_to_excel(ws, START_ROW, "H", R.measurements["Average Friction 100m"], format="0.00")
    write_column_to_excel(ws, START_ROW, "I", R.measurements["Thirds"], format="0.00")

    merge_rows_in_range(len(L), "D", ws, START_ROW, MERGE_RANGE)
    merge_rows_in_range(len(R), "H", ws, START_ROW, MERGE_RANGE)

    merge_columns_into_thirds(len(L), "E", ws, START_ROW)
    merge_columns_into_thirds(len(R), "I", ws, START_ROW)

    center_and_bold_cells(ws, START_ROW, START_COL)

    output_folder_path = Path(output_folder)
    output_file_path = output_folder_path / f"Datos {name}.xlsx"

    wb.save(str(output_file_path))


def write_report_with_chainage(
    L: ASFT_Data, R: ASFT_Data, runway_length: int, starting_point: int, output_folder: str
) -> None:
    validate_attributes(L, R, ["iata", "runway", "numbering", "separation", "equipment", "tyre_type"])
    validate_lengths(L, R)

    L_chainage = L.measurements_with_chainage(runway_length, starting_point)
    R_chainage = R.measurements_with_chainage(runway_length, starting_point)
    L_chainage = L_chainage[L_chainage["Distance"] != 0]
    R_chainage = R_chainage[R_chainage["Distance"] != 0]
    L_chainage = L_chainage.reset_index(drop=True)
    R_chainage = R_chainage.reset_index(drop=True)

    L_chainage["Average Friction 100m"] = friction_interval_mean(L_chainage, "Friction")
    L_chainage["Thirds"] = friction_thirds(L)
    R_chainage["Average Friction 100m"] = friction_interval_mean(R_chainage, "Friction")
    R_chainage["Thirds"] = friction_thirds(R)

    name = get_file_name(L)

    wb, ws = setup_workbook("app/templates/report_template_with_chainage.xlsx", name)

    populate_header_data(L, R, ws)

    write_column_to_excel(ws, START_ROW, "B", L_chainage["Chainage"], format="General")
    write_column_to_excel(ws, START_ROW, "C", L_chainage["Distance"], format="General")
    write_column_to_excel(ws, START_ROW, "D", L_chainage["Friction"], format="0.00")
    write_column_to_excel(ws, START_ROW, "E", L_chainage["Average Friction 100m"], format="0.00")
    write_column_to_excel(ws, START_ROW, "F", L_chainage["Thirds"], format="0.00")
    write_column_to_excel(ws, START_ROW, "G", R_chainage["Friction"], format="0.00")
    write_column_to_excel(ws, START_ROW, "H", R_chainage["Average Friction 100m"], format="0.00")
    write_column_to_excel(ws, START_ROW, "I", R_chainage["Thirds"], format="0.00")

    merge_rows_in_range(len(L), "E", ws, START_ROW, MERGE_RANGE)
    merge_rows_in_range(len(R), "H", ws, START_ROW, MERGE_RANGE)

    merge_columns_into_thirds(len(L), "F", ws, START_ROW)
    merge_columns_into_thirds(len(R), "I", ws, START_ROW)

    center_and_bold_cells(ws, START_ROW, START_COL)

    output_folder_path = Path(output_folder)
    output_file_path = output_folder_path / f"Datos {name}.xlsx"

    wb.save(str(output_file_path))
