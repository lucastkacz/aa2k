import os
import pandas as pd
from typing import Any, List, Optional

from app.models.ASFT_Data import ASFT_Data
from app.utils.calculations import friction_thirds, friction_interval_mean

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment


def write_column(
    data: List[Any],
    start_row: int,
    column: str,
    ws: Worksheet,
    number_format: Optional[str] = None,
) -> None:
    column_index: int = column_index_from_string(column)
    for i, value in enumerate(data, start=start_row):
        if isinstance(value, str):
            try:
                value = float(value)
            except ValueError:
                pass
        cell = ws.cell(row=i, column=column_index)
        cell.value = value
        if number_format:
            cell.number_format = number_format


def write_report(
    L: ASFT_Data,
    R: ASFT_Data,
    estado: str,
    pavimento: str,
    output_folder: str = "output",
) -> None:
    def merge_thirds_column(length: int, col: str, ws: Worksheet, start_row) -> pd.Series:
        third: float = length / 3
        row_A: int = start_row
        row_B: int = round(third) + start_row
        row_C: int = round(2 * third) + start_row
        end_row: int = length + start_row
        ws.merge_cells(f"{col}{row_A}:{col}{row_B - 1}")
        ws.merge_cells(f"{col}{row_B}:{col}{row_C - 1}")
        ws.merge_cells(f"{col}{row_C}:{col}{end_row - 1}")

    def merge_average_row(length: int, col: str, ws: Worksheet, start_row: int, merge_range: int) -> None:
        for row in range(start_row, length + start_row, merge_range):
            ws.merge_cells(f"{col}{row}:{col}{row + merge_range - 1}")

    def format_cells(ws: Worksheet) -> None:
        border = Border(
            left=Side(border_style="medium", color="000000"),
            right=Side(border_style="medium", color="000000"),
            top=Side(border_style="medium", color="000000"),
            bottom=Side(border_style="medium", color="000000"),
        )
        for row in ws.iter_rows(min_row=11, min_col=2):
            if row[0].row == 1:
                continue
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

    L.measurements["Av. Friction 100m"] = friction_interval_mean(L.measurements, "Friction")
    L.measurements["Thirds"] = friction_thirds(L)
    R.measurements["Av. Friction 100m"] = friction_interval_mean(R.measurements, "Friction")
    R.measurements["Thirds"] = friction_thirds(R)

    rows: int = len(L.measurements)

    wb = load_workbook("app/templates/report_template.xlsx")
    ws = wb.active
    ws.title = f"{L.iata}-{L.numbering}-{L.separation}m-{L.date.date()}"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.scale = 95

    # TODO: add data validation to the cells. they must match

    ws["C3"] = L.iata
    ws["E3"] = L.runway
    ws["G3"] = L.numbering
    ws["I3"] = f"{L.separation} m"
    ws["D4"] = L.date.date()
    ws["I5"] = L.equipment[-3:]
    ws["E6"] = int(L.average_speed)
    ws["H6"] = L.tyre_type
    ws["E7"] = estado
    ws["H7"] = pavimento
    ws["E8"] = L.date.time()
    ws["H8"] = R.date.time()

    write_column(L.measurements["Distance"], 11, "B", ws, number_format="General")
    write_column(L.measurements["Friction"], 11, "C", ws, number_format="0.00")
    write_column(L.measurements["Av. Friction 100m"], 11, "D", ws, number_format="0.00")
    write_column(L.measurements["Thirds"], 11, "E", ws, number_format="0.00")
    write_column(R.measurements["Distance"], 11, "F", ws, number_format="General")
    write_column(R.measurements["Friction"], 11, "G", ws, number_format="0.00")
    write_column(R.measurements["Av. Friction 100m"], 11, "H", ws, number_format="0.00")
    write_column(R.measurements["Thirds"], 11, "I", ws, number_format="0.00")

    merge_average_row(rows, "D", ws, 11, 10)
    merge_average_row(rows, "H", ws, 11, 10)

    merge_thirds_column(rows, "E", ws, 11)
    merge_thirds_column(rows, "I", ws, 11)

    format_cells(ws)

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    file_name = f"Datos {L.iata} RWY{L.numbering} {L.separation}m {L.date.date()}.xlsx"
    output_path = os.path.join(output_folder, file_name)
    wb.save(output_path)
